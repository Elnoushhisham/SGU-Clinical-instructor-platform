#!/usr/bin/env python3
"""
publish_data.py — SGU Clinical Roster data publisher
======================================================
Reads the master Excel workbook, filters to April + May 2026,
optionally scrubs PII, and writes roster-data.json ready to
push to GitHub Pages (or any static host).

Usage
-----
  python publish_data.py                            # uses default path below
  python publish_data.py path/to/workbook.xlsx
  python publish_data.py workbook.xlsx --scrub-pii  # remove emails before publishing
  python publish_data.py workbook.xlsx --out custom-name.json

Requirements
------------
  pip install openpyxl
"""

import sys, json, re, argparse, hashlib
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: run  pip install openpyxl")

# ── Configuration ────────────────────────────────────────────────────────────
DEFAULT_XLSX   = "Year2 Master sheet Spring 2026 web app.xlsx"
DEFAULT_OUTPUT = "roster-data.json"
KEEP_MONTHS    = {"2026-04", "2026-05"}          # only April and May 2026

# Roles that mean "not actively occupying" — skip for conflict checks but keep
NON_ACTIVE_ROLES = {
    "absent", "e.absent", "off", "off duty", "tec. failure",
    "technical failure", "internet issue", "other", "test",
    "shadow", "absent all day", "e.absent all day",
}

# ── Helpers ───────────────────────────────────────────────────────────────────
def vstr(v):
    if v is None: return ""
    return str(v).strip()

def norm_name(name):
    name = re.sub(r"\s+", " ", vstr(name)).strip()
    name = re.sub(r"^(Dr\.?|Mr\.?|Mrs\.?|Ms\.?|Prof(?:essor)?\.?)\s+", "", name, flags=re.I)
    return name

def make_id(prefix, *parts):
    raw = "|".join(str(p) for p in parts)
    return prefix + "_" + hashlib.md5(raw.encode()).hexdigest()[:8]

def fmt_date(val):
    """Return YYYY-MM-DD string from openpyxl cell value (datetime or string)."""
    if val is None: return ""
    if hasattr(val, "strftime"): return val.strftime("%Y-%m-%d")
    s = vstr(val)
    # try common formats
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%b-%Y"):
        try:
            from datetime import datetime as dt
            return dt.strptime(s, fmt).strftime("%Y-%m-%d")
        except Exception:
            pass
    return s

def fmt_time(val):
    """Return HH:MM string."""
    if val is None: return ""
    if hasattr(val, "strftime"): return val.strftime("%H:%M")
    s = vstr(val)
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    if m: return f"{int(m.group(1)):02d}:{m.group(2)}"
    # handle "4:05pm" style
    m2 = re.match(r"(\d{1,2}):(\d{2})\s*(am|pm)", s, re.I)
    if m2:
        h, mn = int(m2.group(1)), int(m2.group(2))
        if m2.group(3).lower() == "pm" and h != 12: h += 12
        if m2.group(3).lower() == "am" and h == 12: h = 0
        return f"{h:02d}:{mn:02d}"
    return s

def in_keep_months(date_str):
    return str(date_str or "")[:7] in KEEP_MONTHS

def header_row(ws):
    """Find the first row that looks like a header (has multiple non-empty string cells)."""
    for row in ws.iter_rows(values_only=True):
        non_empty = [c for c in row if c and isinstance(c, str)]
        if len(non_empty) >= 3:
            return [vstr(c).lower().strip() for c in row]
    return []

# ── Sheet parsers ─────────────────────────────────────────────────────────────

def parse_instructors(ws):
    """Parse a department sheet or the 'Clinical Instructor Name' sheet."""
    rows = list(ws.iter_rows(values_only=True))
    instructors = []
    # find header
    hdr_idx = None
    hdrs = []
    for idx, row in enumerate(rows):
        non_empty = [c for c in row if c and isinstance(c, str) and len(vstr(c)) > 1]
        if len(non_empty) >= 2:
            hdrs = [vstr(c).lower() for c in row]
            hdr_idx = idx
            break
    if hdr_idx is None: return instructors

    def col(name_fragments):
        for frag in name_fragments:
            for i, h in enumerate(hdrs):
                if frag in h: return i
        return None

    i_name  = col(["full name", "name"])
    i_first = col(["first"])
    i_last  = col(["last"])
    i_email = col(["email"])
    i_dept  = col(["department"])
    i_slot  = col(["slot", "department slot"])

    for row in rows[hdr_idx+1:]:
        if not any(row): continue
        name = ""
        if i_name is not None: name = vstr(row[i_name]) if i_name < len(row) else ""
        if not name and i_first is not None and i_last is not None:
            f = vstr(row[i_first]) if i_first < len(row) else ""
            l = vstr(row[i_last])  if i_last  < len(row) else ""
            name = (f + " " + l).strip()
        name = norm_name(name)
        if not name or name in ("Not Yet added/Resigned", "#REF!"): continue
        email = vstr(row[i_email]).lower() if i_email is not None and i_email < len(row) else ""
        dept  = vstr(row[i_dept])  if i_dept  is not None and i_dept  < len(row) else ""
        slot  = vstr(row[i_slot])  if i_slot  is not None and i_slot  < len(row) else ""
        inst_id = make_id("inst", email or name)
        instructors.append(dict(
            instructor_id=inst_id, full_name=name, email=email,
            department=dept, department_slot=slot,
            active_status="active", source_sheets=[ws.title],
        ))
    return instructors


def parse_april_may(ws, existing_instructors):
    """
    Parse wide-layout monthly sheet (April or May).
    Row 2 = dates (sparse, forward-fill)
    Row 3 = course
    Row 4 = session name (forward-fill across cohort cols)
    Row 5 = cohort A/B/C/D
    Row 7 = venue (forward-fill)
    Row 10 = start time
    Rows 11+ = one instructor per row; cell value = role at that session/cohort
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 11: return [], []

    activities = []
    assignments = []

    def fwd_fill(row_vals):
        out, last = [], None
        for v in row_vals:
            if v is not None and vstr(v): last = v
            out.append(last)
        return out

    date_row  = fwd_fill(rows[1])   # row 2 (0-indexed: 1)
    course_row = list(rows[2])       # row 3
    name_row  = fwd_fill(rows[3])    # row 4
    cohort_row = list(rows[4])       # row 5
    venue_row = fwd_fill(rows[6])    # row 7
    time_row  = list(rows[9])        # row 10

    inst_name_to_id = {i["full_name"].lower(): i["instructor_id"] for i in existing_instructors}
    # also index by bare name without titles
    for inst in existing_instructors:
        bare = norm_name(inst["full_name"]).lower()
        inst_name_to_id.setdefault(bare, inst["instructor_id"])

    act_cache = {}  # (date, session_name, cohort) → activity_id

    for row in rows[10:]:  # from row 11 onward
        if not any(row): continue
        inst_name = norm_name(vstr(row[0]))
        if not inst_name: continue
        inst_id = inst_name_to_id.get(inst_name.lower())
        if not inst_id:
            # create a provisional instructor record
            inst_id = make_id("inst", inst_name)
            existing_instructors.append(dict(
                instructor_id=inst_id, full_name=inst_name, email="",
                department="", department_slot="",
                active_status="active", source_sheets=[ws.title],
            ))
            inst_name_to_id[inst_name.lower()] = inst_id

        for col_idx in range(1, len(row)):
            cell_val = vstr(row[col_idx]) if col_idx < len(row) else ""
            if not cell_val: continue
            date_str = fmt_date(date_row[col_idx]) if col_idx < len(date_row) else ""
            if not date_str or not in_keep_months(date_str): continue

            session_name = vstr(name_row[col_idx]) if col_idx < len(name_row) else ""
            cohort       = vstr(cohort_row[col_idx]) if col_idx < len(cohort_row) else ""
            venue        = vstr(venue_row[col_idx]) if col_idx < len(venue_row) else ""
            course       = vstr(course_row[col_idx]) if col_idx < len(course_row) else ""
            start_time   = fmt_time(time_row[col_idx]) if col_idx < len(time_row) else ""
            role         = cell_val

            # Determine activity_type from session name
            act_type = "SG"
            if re.search(r"\bITI\b", session_name, re.I):   act_type = "ITI"
            elif re.search(r"\bBLS\b", session_name, re.I): act_type = "BLS"
            elif re.search(r"\bLecture\b", session_name, re.I): act_type = "Lecture"
            elif re.search(r"\bSGP\b", session_name, re.I): act_type = "SGP"
            elif re.search(r"\bReview\b", session_name, re.I): act_type = "Review"

            act_key = (date_str, session_name, cohort)
            if act_key not in act_cache:
                act_id = make_id("act", date_str, session_name, cohort)
                act_cache[act_key] = act_id
                activities.append(dict(
                    activity_id=act_id, course=course,
                    activity_name=session_name + (" ("+cohort+")" if cohort else ""),
                    activity_type=act_type, date=date_str,
                    start_time=start_time, end_time="",
                    location=venue, discipline="", description="",
                    required_number_of_instructors=1, notes="",
                    worksheet_source=ws.title,
                ))
            act_id = act_cache[act_key]

            asgn_id = make_id("asgn", inst_id, act_id)
            assignments.append(dict(
                assignment_id=asgn_id, activity_id=act_id,
                instructor_id=inst_id, date=date_str,
                start_time=start_time, end_time="",
                location=venue, role=role,
                is_active=(role.lower().strip() not in NON_ACTIVE_ROLES),
                worksheet_source=ws.title,
            ))
    return activities, assignments


def parse_review_sessions(ws):
    """Parse the 'Review sessions Timetable' sheet."""
    rows = list(ws.iter_rows(values_only=True))
    reviews = []
    hdr_idx, hdrs = None, []
    for idx, row in enumerate(rows):
        non_empty = [c for c in row if c and isinstance(c, str)]
        if len(non_empty) >= 4:
            hdrs = [vstr(c).lower() for c in row]
            hdr_idx = idx; break
    if hdr_idx is None: return reviews

    def col(frags):
        for frag in frags:
            for i, h in enumerate(hdrs):
                if frag in h: return i
        return None

    i_date  = col(["date"])
    i_start = col(["time", "start"])
    i_dur   = col(["duration"])
    i_sg    = col(["sg#", "sg ", "sg_code", "sg code"])
    i_disc  = col(["discipline"])
    i_venue = col(["venue", "location"])
    i_zoom  = col(["zoom"])
    i_mtg   = col(["meeting id", "meeting"])
    i_pass  = col(["passcode", "password"])
    i_course= col(["course"])
    i_module= col(["module"])
    i_notes = col(["notes"])

    def get(row, i):
        return vstr(row[i]) if i is not None and i < len(row) else ""

    for row in rows[hdr_idx+1:]:
        if not any(row): continue
        date_str = fmt_date(row[i_date]) if i_date is not None and i_date < len(row) else ""
        if not date_str or not in_keep_months(date_str): continue
        rev_id = make_id("rev", date_str, get(row, i_sg), get(row, i_start))
        reviews.append(dict(
            review_id=rev_id,
            linked_activity_id=None,
            course=get(row, i_course), module=get(row, i_module),
            discipline=get(row, i_disc),
            review_date=date_str,
            review_start_time=fmt_time(row[i_start]) if i_start is not None and i_start < len(row) else "",
            review_end_time="", duration=get(row, i_dur),
            review_location=get(row, i_venue),
            zoom_link=get(row, i_zoom), meeting_id=get(row, i_mtg), passcode=get(row, i_pass),
            sg_code=get(row, i_sg), notes=get(row, i_notes),
        ))
    return reviews


# ── Main ──────────────────────────────────────────────────────────────────────

def build_dataset(xlsx_path, scrub_pii=False):
    print(f"Opening: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")

    instructors  = []
    master_sched = []
    assignments  = []
    reviews      = []

    DEPT_SHEETS = {"pathophysiology", "pathology", "clinical skills", "micro pharm",
                   "clinical instructor name", "deparmtment for the link",
                   "contact infromation"}
    MONTH_SHEETS = {"april", "may"}
    REVIEW_SHEETS = {"review sessions timetable", "review session attendance"}

    # Parse instructor/department sheets first
    for sheet_name in wb.sheetnames:
        norm = sheet_name.lower().strip()
        if any(d in norm for d in DEPT_SHEETS):
            insts = parse_instructors(wb[sheet_name])
            instructors.extend(insts)
            print(f"  {sheet_name}: {len(insts)} instructors")

    # Deduplicate instructors (by email then by name)
    seen_keys = {}
    deduped = []
    for inst in instructors:
        key = inst["email"] or inst["full_name"].lower()
        if key not in seen_keys:
            seen_keys[key] = len(deduped)
            deduped.append(inst)
        else:
            # Merge source_sheets
            existing = deduped[seen_keys[key]]
            for s in inst["source_sheets"]:
                if s not in existing["source_sheets"]:
                    existing["source_sheets"].append(s)
            # Fill in missing fields
            if not existing["email"] and inst["email"]:
                existing["email"] = inst["email"]
            if not existing["department"] and inst["department"]:
                existing["department"] = inst["department"]
    instructors = deduped
    print(f"  Instructors after dedup: {len(instructors)}")

    # Parse month sheets
    for sheet_name in wb.sheetnames:
        if sheet_name.lower().strip() in MONTH_SHEETS:
            acts, asgns = parse_april_may(wb[sheet_name], instructors)
            master_sched.extend(acts)
            assignments.extend(asgns)
            print(f"  {sheet_name}: {len(acts)} activities, {len(asgns)} assignments")

    # Parse review sessions
    for sheet_name in wb.sheetnames:
        norm = sheet_name.lower().strip()
        if "review sessions timetable" in norm:
            revs = parse_review_sessions(wb[sheet_name])
            reviews.extend(revs)
            print(f"  {sheet_name}: {len(revs)} reviews")

    # Link reviews to SG activities by sg_code substring match
    for rev in reviews:
        sg = rev["sg_code"].lower()
        if not sg: continue
        for act in master_sched:
            if sg in act["activity_name"].lower() or act["activity_name"].lower() in sg:
                rev["linked_activity_id"] = act["activity_id"]
                break

    # Scrub PII if requested
    if scrub_pii:
        for inst in instructors:
            inst["email"] = ""
        print("  PII scrubbed (emails removed)")

    dataset = dict(
        meta=dict(
            source_file=str(xlsx_path),
            parsed_at=datetime.utcnow().isoformat() + "Z",
            sheets_seen=wb.sheetnames,
            keep_months=sorted(KEEP_MONTHS),
            scrubbed_pii=scrub_pii,
        ),
        instructors=instructors,
        master_schedule=master_sched,
        review_sessions=reviews,
        assignments=assignments,
        review_attendance=[],
        locations=sorted({a["location"] for a in master_sched if a["location"]}),
        courses=sorted({a["course"] for a in master_sched if a["course"]}),
        leave_requests=[],
    )
    return dataset


def main():
    parser = argparse.ArgumentParser(description="Publish SGU roster data to JSON")
    parser.add_argument("xlsx", nargs="?", default=DEFAULT_XLSX, help="Path to .xlsx file")
    parser.add_argument("--out",       default=DEFAULT_OUTPUT, help="Output JSON file")
    parser.add_argument("--scrub-pii", action="store_true",    help="Remove email addresses before publishing")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        sys.exit(f"File not found: {xlsx_path}")

    dataset = build_dataset(xlsx_path, scrub_pii=args.scrub_pii)

    out_path = Path(args.out)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(dataset, f, separators=(",", ":"), ensure_ascii=False)

    size_kb = out_path.stat().st_size // 1024
    print(f"\nWrote {out_path}  ({size_kb} KB)")
    print(f"  {len(dataset['instructors'])} instructors")
    print(f"  {len(dataset['master_schedule'])} activities  (April + May 2026)")
    print(f"  {len(dataset['assignments'])} assignments")
    print(f"  {len(dataset['review_sessions'])} review sessions")
    print(f"\nNext steps:")
    print(f"  git add {out_path} && git commit -m 'Update roster data' && git push")
    print(f"  Then in the app: Coordinator → Import → Fetch published JSON URL")


if __name__ == "__main__":
    main()
